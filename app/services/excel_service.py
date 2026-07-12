from openpyxl import load_workbook

from app.core.config import settings

import warnings

warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl"
)

class ExcelService:

    def __init__(self):
        self.file = settings.EXCEL_FILE

    def read(self):

        wb = load_workbook(
            self.file,
            data_only=True
        )

        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        headers = [str(x).strip() if x else "" for x in rows[1]]

        data = []

        for row in rows[2:]:

            obj = {}

            for h, v in zip(headers, row):

                obj[h] = "" if v is None else str(v)

            data.append(obj)

        return data